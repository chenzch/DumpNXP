import openpyxl
from openpyxl.styles import Alignment

from IParseItem import IParseItem


class BitItem(IParseItem):
    def __init__(self, name: str, offset: int, true_meaning: str, false_meaning: str):
        self.name = name
        self.offset = offset
        self.true_meaning = true_meaning
        self.false_meaning = false_meaning

    def ParseItem(self, value: int, sheet: openpyxl.worksheet.worksheet.Worksheet, start_row: int, start_column: int) -> int:
        # 设置对齐方式
        alignment = Alignment(horizontal='right', vertical='center')

        result = bool(value & (1 << self.offset))
        text = self.true_meaning if result else self.false_meaning
        sheet.cell(row=start_row, column=start_column, value=self.name)
        sheet.cell(row=start_row, column=start_column).alignment = alignment
        sheet.cell(row=start_row, column=start_column + 1, value=(value >> self.offset) & 1)
        sheet.cell(row=start_row, column=start_column + 2, value=text)
        return 1
