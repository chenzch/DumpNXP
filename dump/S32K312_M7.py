import sys
import re
import openpyxl

from openpyxl.styles import Alignment




def ReadFile(filename: str) -> list:
    result = []
    pattern = r'^~\"V/([0-9A-Fa-f]{8})/([0-9A-Fa-f]+)\\n\"'

    try:
        with open(filename, 'r') as file:
            for line in file:
                line = line.strip()
                match = re.match(pattern, line)
                if match:
                    hex1, hex2 = match.groups()
                    result.append((int(hex1, 16), int(hex2, 16)))  # 将两个 uint32 数字作为元组追加
    except FileNotFoundError:
        print(f"错误: 找不到文件 '{filename}'")
    except IOError:
        print(f"错误: 无法读取文件 '{filename}'")

    return result


def Parse(input_filename: str, excel_filename: str):

    # 创建新的Excel工作簿和工作表
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    result = ReadFile(input_filename)
    for value1, value2 in result:
        print("~\"V/{:X}/{:08X}\\n\"".format(value1, value2))

    # for module in device.GetModuleList():
    #     # 获取模块名称并创建新的工作表
    #     module_sheet_name = module.GetModuleName()
    #     module_sheet = workbook.create_sheet(title=module_sheet_name)

    #     # 调用UpdateSheet方法更新工作表
    #     module.UpdateSheet(result, module_sheet)

    if len(workbook.sheetnames) > 1:
        for sheet in workbook.sheetnames:
            module_sheet = workbook[sheet]
            # 写入表头
            module_sheet['A1'] = "名称"
            module_sheet['B1'] = "值"
            module_sheet['C1'] = "意义"

            # 设置对齐方式
            alignment = Alignment(horizontal='center', vertical='center')
            for row in module_sheet.iter_rows(min_row=1, max_col=3):
                for cell in row:
                    cell.alignment = alignment

            # 自动调整列宽度
            for column in ['A', 'B', 'C']:
                module_sheet.column_dimensions[column].width = max(module_sheet.column_dimensions[column].width, max(len(str(cell.value)) for cell in module_sheet[column])) + 2
        workbook.save(excel_filename)


def main():
    if len(sys.argv) != 2:
        print("使用方法: python S32K312_M7.py <输入文件名>")
        sys.exit(1)

    input_filename = sys.argv[1]
    excel_filename = input_filename.rsplit('.', 1)[0] + '.xlsx'

    Parse(input_filename, excel_filename)


if __name__ == "__main__":
    main()
