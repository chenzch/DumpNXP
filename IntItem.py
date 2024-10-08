from typing import List

import openpyxl
from openpyxl.styles import Alignment

from IParseItem import IParseItem


class IntItem(IParseItem):
    def __init__(self, name: str, offset: int, width: int, enumStr: List = None, description: str = None):
        self.name = name
        self.offset = offset
        self.width = width
        self.enumStr = enumStr
        self.description = description

    def ParseItem(self, value: int, sheet: openpyxl.worksheet.worksheet.Worksheet, start_row: int, start_column: int) -> int:
        # 设置对齐方式
        alignment = Alignment(horizontal='right', vertical='center')

        value = (value >> self.offset) & ((1 << self.width) - 1)

        sheet.cell(row=start_row, column=start_column, value=self.name)
        sheet.cell(row=start_row, column=start_column).alignment = alignment
        sheet.cell(row=start_row, column=start_column + 1, value=value)
        if self.enumStr is not None:
            sheet.cell(row=start_row, column=start_column + 2, value=self.enumStr[value])
        elif self.description is not None:
            sheet.cell(row=start_row, column=start_column + 2, value=self.description)
        return 1
