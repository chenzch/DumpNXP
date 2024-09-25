from typing import List, Tuple

import openpyxl

from IModule import IModule
from IntItem import IntItem
from Parse import ProcessArray


class SBAF(IModule):
    def __init__(self, name: str, address: int) -> None:
        self.Name = name
        self.Address = address

    def GetModuleName(self) -> str:
        return self.Name

    def UpdateSheet(self, data: List[Tuple[str, str]], sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        row = 1
        sbaf_version_base = self.Address
        for item in data:
            row += 1
            addr = int(item[0], 16)
            value = int(item[1], 16)
            if (addr == sbaf_version_base):
                sheet.cell(row=row, column=1, value="SBAF_VERSION")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name="SBAF_FW_TYPE", offset=16, width=16),
                    IntItem(name="SBAF_SOC_TYPE_ID", offset=8, width=8, enumStr=[
                        "", "", "", "", "", "S32K344, S32K324, S32K314", "", "", "", "", "", "S32K310",
                        "S32K311, S32K341", "S32K312, S32K322, S32K342", "S32K358, S32K348, S32K338, S32K328",
                        "S32K396, S32K376, S32K394, S32K374", "S32K388", "S32K389"
                    ]),
                    IntItem(name="SBAF_IN_ABSWAP", offset=0, width=8),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == sbaf_version_base + 0x04):
                sheet.cell(row=row, column=1, value="SBAF_VERSION")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name="SBAF_RC_NUMBER", offset=24, width=8),
                    IntItem(name="SBAF_INCREMENTAL_NUMBER", offset=16, width=8),
                    IntItem(name="SBAF_BASELINE_NUMBER", offset=8, width=8),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            else:
                row -= 1
