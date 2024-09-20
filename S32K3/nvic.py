from typing import List, Tuple
from enum import Enum
import openpyxl

from BitItem import BitItem
from IModule import IModule
from IntItem import IntItem
from Parse import ProcessArray


class NVICType(Enum):
    S32K3 = 1
    S32K3E = 2

class NVIC(IModule):
    def __init__(self, name: str, address: int, Type: NVICType):
        self.Name = name
        self.Address = address
        if (Type == NVICType.S32K3E):
            self.NameList = [
                "INT0",
                "INT1",
                "INT2",
                "INT3",
                "eDMA0_DMATCD_CH0_CH1",
                "eDMA0_DMATCD_CH2_CH3",
                "eDMA0_DMATCD_CH4_CH5",
                "eDMA0_DMATCD_CH6_CH7",
                "eDMA0_DMATCD_CH8_CH9",
                "eDMA0_DMATCD_CH10_CH11",
                "eDMA0_DMATCD_CH12_CH13",
                "eDMA0_DMATCD_CH14_CH15",
                "eDMA0_DMATCD_CH16_CH17",
                "eDMA0_DMATCD_CH18_CH19",
                "eDMA0_DMATCD_CH20_CH21",
                "eDMA0_DMATCD_CH22_CH23",
                "eDMA0_DMATCD_CH24_CH25",
                "eDMA0_DMATCD_CH26_CH27",
                "eDMA0_DMATCD_CH28_CH29",
                "eDMA0_DMATCD_CH30_CH31",
                "eDMA1_DMATCD_CH0_CH1",
                "eDMA1_DMATCD_CH2_CH3",
                "eDMA1_DMATCD_CH4_CH5",
                "eDMA1_DMATCD_CH6_CH7",
                "eDMA1_DMATCD_CH8_CH9",
                "eDMA1_DMATCD_CH10_CH11",
                "eDMA1_DMATCD_CH12_CH13",
                "eDMA1_DMATCD_CH14_CH15",
                "eDMA1_DMATCD_CH16_CH17",
                "eDMA1_DMATCD_CH18_CH19",
                "eDMA1_DMATCD_CH20_CH21",
                "eDMA1_DMATCD_CH22_CH23",
                "eDMA1_DMATCD_CH24_CH25",
                "eDMA1_DMATCD_CH26_CH27",
                "eDMA1_DMATCD_CH28_CH29",
                "eDMA1_DMATCD_CH30_CH31",
                "ERM_0",
                "ERM_1",
                "MCM",
                "STM0",
                "STM1",
                "STM2",
                "SWT0",
                "SWT1",
                "SWT2",
                "CTI0",
                "CTI1",
                "CTI2",
                "FLASH_0",
                "FLASH_1",
                "FLASH_2",
                "RGM",
                "PMC",
                "SIUL_0",
                "SIUL_1",
                "SIUL_2",
                "SIUL_3",
                "eTPU",
                "eTPU_A_CH0",
                "eTPU_A_CH1",
                "eTPU_A_CH2",
                "EMIOS0_0",
                "EMIOS0_1",
                "EMIOS0_2",
                "EMIOS0_3",
                "EMIOS0_4",
                "EMIOS0_5",
                "SIPI1_CH1",
                "SIPI1_CH2",
                "SIPI1_CH3",
                "SIPI1_CH4",
                "SIPI1_Int1",
                "SIPI1_Int2",
                "SIPI1_Int3",
                "DIGRF_Tx_int",
                "DIGRF_Tx_exc",
                "DIGRF_Rx_int",
                "DIGRF_Rx_exc",
                "DIGRF_ICLC_Rx",
                "eTPU_A_CH3",
                "eTPU_A_CH4",
                "eTPU_A_CH5",
                "eTPU_A_CH6",
                "WKPU",
                "CMU0",
                "CMU1",
                "CMU2",
                "BCTU0",
                "BCTU1",
                "eTPU_A_CH7",
                "eTPU_A_CH8",
                "eTPU_A_CH9",
                "LCU0",
                "LCU1",
                "eTPU_A_CH10",
                "eTPU_A_CH11",
                "PIT0",
                "PIT1",
                "PIT2",
                "eTPU_A_CH12",
                "eTPU_A_CH13",
                "eTPU_A_CH14",
                "RTC",
                "eTPU_A_CH15",
                "eTPU_A_CH16",
                "EMAC_0",
                "EMAC_1",
                "EMAC_2",
                "EMAC_3",
                "FlexCAN0_0",
                "FlexCAN0_1",
                "FlexCAN0_2",
                "FlexCAN0_3",
                "FlexCAN1_0",
                "FlexCAN1_1",
                "FlexCAN1_2",
                "FlexCAN2_0",
                "FlexCAN2_1",
                "FlexCAN2_2",
                "FlexCAN3_0",
                "FlexCAN3_1",
                "FlexCAN4_0",
                "FlexCAN4_1",
                "FlexCAN5_0",
                "FlexCAN5_1",
                "eTPU_A_CH17",
                "eTPU_A_CH18",
                "eTPU_A_CH19",
                "eTPU_A_CH20",
                "FlexCAN1_3",
                "FlexCAN2_3",
                "FlexCAN3_2",
                "FlexCAN4_2",
                "FlexCAN5_2",
                "eTPU_A_CH21",
                "eTPU_A_CH22",
                "eTPU_A_CH23",
                "SWG_0",
                "eTPU_A_CH24_31",
                "FLEXIO",
                "eTPU_B_CH0",
                "LPUART0",
                "LPUART1",
                "LPUART2",
                "LPUART3",
                "LPUART4",
                "LPUART5",
                "eTPU_B_CH1",
                "eTPU_B_CH2",
                "eTPU_B_CH3",
                "eTPU_B_CH4",
                "eTPU_B_CH5",
                "eTPU_B_CH6",
                "eTPU_B_CH7",
                "eTPU_B_CH8",
                "eTPU_B_CH9",
                "eTPU_B_CH10",
                "eTPU_B_CH11",
                "eTPU_B_CH12",
                "eTPU_B_CH13",
                "",
                "LPI2C0",
                "LPI2C1",
                "",
                "DSPI0",
                "LPSPI0",
                "LPSPI1",
                "LPSPI2",
                "LPSPI3",
                "LPSPI4",
                "LPSPI5",
                "eTPU_B_CH14",
                "eTPU_B_CH15",
                "QSPI",
                "eTPU_B_CH16",
                "eTPU_B_CH17",
                "eFlexPWM0",
                "eFlexPWM1",
                "JDC",
                "SDADC",
                "ADC0",
                "ADC1",
                "ADC2",
                "LPCMP0",
                "LPCMP1",
                "ADC3",
                "ADC4",
                "ADC5",
                "ADC6",
                "FCCU_0",
                "FCCU_1",
                "STCU_LBIST_MBIST",
                "HSE_MU0_TX",
                "HSE_MU0_RX",
                "HSE_MU0_ORED",
                "HSE_MU1_TX",
                "HSE_MU1_RX",
                "HSE_MU1_ORED",
                "eTPU_B_CH18",
                "eTPU_B_CH19",
                "eTPU_B_CH20",
                "eTPU_B_CH21",
                "MU_A_0",
                "MU_A_1",
                "MU_A_2",
                "MU_B_0",
                "MU_B_1",
                "MU_B_2",
                "",
                "",
                "",
                "",
                "Soc_RSB",
                "",
                "eFlexPWM1_1",
                "",
                "XRDC",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "CoolFlux1",
                "CoolFlux2",
                "CoolFlux3",
                "CoolFlux4",
                "CoolFlux5",
                "eTPU_B_CH22",
                "eTPU_B_CH23",
                "eTPU_B_CH24",
                "SWG1",
                "eTPU_B_CH25_31",
                "eFlexPWM0_0",
                "eFlexPWM0_1",
                "eFlexPWM0_2",
                "eFlexPWM0_3",
                "eFlexPWM1_2",
                "eFlexPWM1_3"
            ]
        else:
            self.NameList = [
                "INT0",
                "INT1",
                "INT2",
                "INT3",
                "DMATCD0",
                "DMATCD1",
                "DMATCD2",
                "DMATCD3",
                "DMATCD4",
                "DMATCD5",
                "DMATCD6",
                "DMATCD7",
                "DMATCD8",
                "DMATCD9",
                "DMATCD10",
                "DMATCD11",
                "DMATCD12",
                "DMATCD13",
                "DMATCD14",
                "DMATCD15",
                "DMATCD16",
                "DMATCD17",
                "DMATCD18",
                "DMATCD19",
                "DMATCD20",
                "DMATCD21",
                "DMATCD22",
                "DMATCD23",
                "DMATCD24",
                "DMATCD25",
                "DMATCD26",
                "DMATCD27",
                "DMATCD28",
                "DMATCD29",
                "DMATCD30",
                "DMATCD31",
                "ERM_0",
                "ERM_1",
                "MCM",
                "STM0",
                "STM1",
                "STM2",
                "SWT0",
                "SWT1",
                "SWT2",
                "CTI0",
                "CTI1",
                "",
                "FLASH_0",
                "FLASH_1",
                "FLASH_2",
                "RGM",
                "PMC",
                "SIUL_0",
                "SIUL_1",
                "SIUL_2",
                "SIUL_3",
                "STM3",
                "SWT3",
                "",
                "",
                "EMIOS0_0",
                "EMIOS0_1",
                "EMIOS0_2",
                "EMIOS0_3",
                "EMIOS0_4",
                "EMIOS0_5",
                "",
                "",
                "EMIOS1_0",
                "EMIOS1_1",
                "EMIOS1_2",
                "EMIOS1_3",
                "EMIOS1_4",
                "EMIOS1_5",
                "",
                "",
                "EMIOS2_0",
                "EMIOS2_1",
                "EMIOS2_2",
                "EMIOS2_3",
                "EMIOS2_4",
                "EMIOS2_5",
                "WKPU",
                "CMU0",
                "CMU1",
                "CMU2",
                "BCTU",
                "AESAPP0",
                "AES_APP1",
                "AES_APP2",
                "AES_APP3",
                "LCU0",
                "LCU1",
                "AES_APP4",
                "AES_APP5",
                "PIT0",
                "PIT1",
                "PIT2",
                "PIT3",
                "AES_APP6",
                "AES_APP7",
                "RTC",
                "",
                "",
                "",
                "",
                "",
                "",
                "FlexCAN0_0",
                "FlexCAN0_1",
                "FlexCAN0_2",
                "FlexCAN0_3",
                "FlexCAN1_0",
                "FlexCAN1_1",
                "FlexCAN1_2",
                "FlexCAN2_0",
                "FlexCAN2_1",
                "FlexCAN2_2",
                "FlexCAN3_0",
                "FlexCAN3_1",
                "FlexCAN4_0",
                "FlexCAN4_1",
                "FlexCAN5_0",
                "FlexCAN5_1",
                "FlexCAN6_0",
                "FlexCAN6_1",
                "FlexCAN7_0",
                "FlexCAN7_1",
                "FlexCAN1_3",
                "FlexCAN2_3",
                "FlexCAN3_2",
                "FlexCAN4_2",
                "FlexCAN5_2",
                "FlexCAN6_2",
                "FlexCAN7_2",
                "MU4_A_0",
                "MU4_A_1",
                "MU4_A_2",
                "FLEXIO",
                "",
                "LPUART0",
                "LPUART1",
                "LPUART2",
                "LPUART3",
                "LPUART4",
                "LPUART5",
                "LPUART6",
                "LPUART7",
                "LPUART8",
                "LPUART9",
                "LPUART10",
                "LPUART11",
                "LPUART12",
                "LPUART13",
                "LPUART14",
                "LPUART15",
                "MU4_B_0",
                "MU4_B_1",
                "MU4_B_2",
                "",
                "LPI2C0",
                "LPI2C1",
                "",
                "",
                "LPSPI0",
                "LPSPI1",
                "LPSPI2",
                "LPSPI3",
                "LPSPI4",
                "LPSPI5",
                "GMAC1_Common",
                "GMAC1_CH0_TX",
                "QSPI",
                "SAI0",
                "SAI1",
                "",
                "",
                "JDC",
                "",
                "ADC0",
                "ADC1",
                "ADC2",
                "LPCMP0",
                "LPCMP1",
                "LPCMP2",
                "MU3_A_0",
                "MU3_A_1",
                "MU3_A_2",
                "FCCU_0",
                "FCCU_1",
                "STCU_LBIST_MBIST",
                "HSE_MU0_TX",
                "HSE_MU0_RX",
                "HSE_MU0_ORED",
                "HSE_MU1_TX",
                "HSE_MU1_RX",
                "HSE_MU1_ORED",
                "",
                "MU3_B_0",
                "MU3_B_1",
                "MU3_B_2",
                "MU_A_0",
                "MU_A_1",
                "MU_A_2",
                "MU_B_0",
                "MU_B_1",
                "MU_B_2",
                "",
                "",
                "",
                "",
                "SoC",
                "Core_clk",
                "SoC2",
                "SoC3",
                "XRDC",
                "SoC4",
                "",
                "SoC5",
                "",
                "",
                "",
                "",
                "GMAC0_Common",
                "GMAC0_CH0_TX",
                "GMAC0_CH0_RX",
                "GMAC0_CH1_TX",
                "GMAC0_CH1_RX",
                "GMAC0_CH2_TX",
                "GMAC0_CH2_RX",
                "GMAC0_SIC",
                "GMAC0_SIUC",
                "GMAC1_CH0_RX",
                "GMAC1_CH1_TX",
                "GMAC1_CH1_RX",
                "GMAC1_CH2_TX",
                "GMAC1_CH2_RX",
                "GMAC1_SIC",
                "GMAC1_SIUC"
            ]

    def GetModuleName(self) -> str:
        return self.Name

    def UpdateSheet(self, data: List[Tuple[str, str]], sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        row = 1
        scb_base = self.Address
        for item in data:
            row += 1
            addr = int(item[0], 16)
            value = int(item[1], 16)
            if (addr == scb_base + 0x100):
                sheet.cell(row=row, column=1, value="ISER0")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[0], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[1], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[2], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[3], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[4], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[5], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[6], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[7], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[8], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[9], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[10], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[11], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[12], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[13], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[14], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[15], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[16], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[17], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[18], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[19], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[20], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[21], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[22], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[23], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[24], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[25], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[26], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[27], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[28], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[29], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[30], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[31], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x104):
                sheet.cell(row=row, column=1, value="ISER1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[32], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[33], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[34], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[35], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[36], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[37], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[38], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[39], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[40], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[41], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[42], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[43], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[44], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[45], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[46], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[47], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[48], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[49], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[50], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[51], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[52], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[53], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[54], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[55], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[56], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[57], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[58], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[59], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[60], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[61], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[62], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[63], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x108):
                sheet.cell(row=row, column=1, value="ISER2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[64], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[65], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[66], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[67], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[68], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[69], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[70], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[71], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[72], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[73], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[74], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[75], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[76], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[77], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[78], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[79], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[80], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[81], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[82], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[83], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[84], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[85], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[86], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[87], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[88], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[89], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[90], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[91], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[92], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[93], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[94], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[95], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x10C):
                sheet.cell(row=row, column=1, value="ISER3")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[96], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[97], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[98], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[99], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[100], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[101], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[102], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[103], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[104], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[105], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[106], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[107], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[108], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[109], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[110], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[111], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[112], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[113], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[114], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[115], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[116], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[117], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[118], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[119], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[120], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[121], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[122], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[123], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[124], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[125], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[126], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[127], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x110):
                sheet.cell(row=row, column=1, value="ISER4")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[128], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[129], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[130], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[131], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[132], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[133], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[134], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[135], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[136], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[137], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[138], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[139], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[140], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[141], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[142], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[143], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[144], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[145], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[146], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[147], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[148], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[149], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[150], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[151], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[152], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[153], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[154], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[155], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[156], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[157], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[158], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[159], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x114):
                sheet.cell(row=row, column=1, value="ISER5")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[160], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[161], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[162], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[163], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[164], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[165], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[166], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[167], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[168], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[169], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[170], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[171], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[172], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[173], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[174], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[175], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[176], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[177], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[178], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[179], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[180], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[181], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[182], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[183], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[184], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[185], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[186], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[187], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[188], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[189], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[190], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[191], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x118):
                sheet.cell(row=row, column=1, value="ISER6")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[192], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[193], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[194], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[195], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[196], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[197], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[198], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[199], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[200], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[201], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[202], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[203], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[204], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[205], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[206], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[207], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[208], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[209], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[210], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[211], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[212], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[213], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[214], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[215], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[216], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[217], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[218], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[219], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[220], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[221], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[222], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[223], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x11C):
                sheet.cell(row=row, column=1, value="ISER7")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[224], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[225], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[226], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[227], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[228], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[229], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[230], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[231], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[232], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[233], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[234], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[235], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[236], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[237], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[238], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[239], offset=15, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x180):
                sheet.cell(row=row, column=1, value="ICER0")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[0], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[1], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[2], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[3], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[4], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[5], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[6], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[7], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[8], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[9], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[10], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[11], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[12], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[13], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[14], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[15], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[16], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[17], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[18], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[19], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[20], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[21], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[22], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[23], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[24], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[25], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[26], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[27], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[28], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[29], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[30], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[31], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x184):
                sheet.cell(row=row, column=1, value="ICER1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[32], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[33], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[34], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[35], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[36], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[37], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[38], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[39], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[40], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[41], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[42], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[43], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[44], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[45], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[46], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[47], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[48], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[49], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[50], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[51], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[52], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[53], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[54], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[55], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[56], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[57], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[58], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[59], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[60], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[61], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[62], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[63], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x188):
                sheet.cell(row=row, column=1, value="ICER2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[64], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[65], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[66], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[67], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[68], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[69], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[70], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[71], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[72], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[73], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[74], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[75], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[76], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[77], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[78], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[79], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[80], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[81], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[82], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[83], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[84], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[85], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[86], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[87], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[88], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[89], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[90], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[91], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[92], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[93], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[94], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[95], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x18C):
                sheet.cell(row=row, column=1, value="ICER3")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[96], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[97], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[98], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[99], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[100], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[101], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[102], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[103], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[104], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[105], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[106], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[107], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[108], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[109], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[110], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[111], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[112], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[113], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[114], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[115], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[116], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[117], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[118], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[119], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[120], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[121], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[122], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[123], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[124], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[125], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[126], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[127], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x190):
                sheet.cell(row=row, column=1, value="ICER4")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[128], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[129], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[130], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[131], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[132], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[133], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[134], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[135], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[136], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[137], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[138], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[139], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[140], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[141], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[142], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[143], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[144], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[145], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[146], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[147], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[148], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[149], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[150], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[151], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[152], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[153], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[154], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[155], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[156], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[157], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[158], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[159], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x194):
                sheet.cell(row=row, column=1, value="ICER5")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[160], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[161], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[162], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[163], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[164], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[165], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[166], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[167], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[168], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[169], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[170], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[171], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[172], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[173], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[174], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[175], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[176], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[177], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[178], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[179], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[180], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[181], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[182], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[183], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[184], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[185], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[186], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[187], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[188], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[189], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[190], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[191], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x198):
                sheet.cell(row=row, column=1, value="ICER6")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[192], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[193], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[194], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[195], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[196], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[197], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[198], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[199], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[200], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[201], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[202], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[203], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[204], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[205], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[206], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[207], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[208], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[209], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[210], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[211], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[212], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[213], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[214], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[215], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[216], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[217], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[218], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[219], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[220], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[221], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[222], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[223], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x19C):
                sheet.cell(row=row, column=1, value="ICER7")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[224], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[225], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[226], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[227], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[228], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[229], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[230], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[231], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[232], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[233], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[234], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[235], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[236], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[237], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[238], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[239], offset=15, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x200):
                sheet.cell(row=row, column=1, value="ISPR0")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[0], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[1], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[2], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[3], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[4], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[5], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[6], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[7], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[8], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[9], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[10], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[11], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[12], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[13], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[14], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[15], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[16], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[17], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[18], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[19], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[20], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[21], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[22], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[23], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[24], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[25], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[26], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[27], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[28], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[29], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[30], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[31], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x204):
                sheet.cell(row=row, column=1, value="ISPR1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[32], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[33], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[34], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[35], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[36], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[37], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[38], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[39], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[40], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[41], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[42], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[43], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[44], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[45], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[46], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[47], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[48], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[49], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[50], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[51], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[52], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[53], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[54], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[55], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[56], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[57], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[58], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[59], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[60], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[61], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[62], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[63], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x208):
                sheet.cell(row=row, column=1, value="ISPR2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[64], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[65], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[66], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[67], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[68], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[69], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[70], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[71], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[72], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[73], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[74], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[75], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[76], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[77], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[78], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[79], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[80], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[81], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[82], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[83], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[84], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[85], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[86], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[87], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[88], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[89], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[90], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[91], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[92], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[93], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[94], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[95], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x20C):
                sheet.cell(row=row, column=1, value="ISPR3")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[96], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[97], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[98], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[99], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[100], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[101], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[102], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[103], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[104], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[105], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[106], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[107], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[108], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[109], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[110], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[111], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[112], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[113], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[114], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[115], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[116], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[117], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[118], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[119], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[120], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[121], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[122], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[123], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[124], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[125], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[126], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[127], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x210):
                sheet.cell(row=row, column=1, value="ISPR4")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[128], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[129], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[130], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[131], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[132], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[133], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[134], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[135], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[136], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[137], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[138], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[139], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[140], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[141], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[142], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[143], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[144], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[145], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[146], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[147], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[148], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[149], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[150], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[151], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[152], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[153], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[154], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[155], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[156], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[157], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[158], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[159], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x214):
                sheet.cell(row=row, column=1, value="ISPR5")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[160], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[161], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[162], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[163], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[164], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[165], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[166], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[167], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[168], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[169], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[170], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[171], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[172], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[173], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[174], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[175], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[176], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[177], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[178], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[179], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[180], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[181], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[182], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[183], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[184], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[185], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[186], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[187], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[188], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[189], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[190], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[191], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x218):
                sheet.cell(row=row, column=1, value="ISPR6")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[192], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[193], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[194], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[195], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[196], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[197], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[198], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[199], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[200], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[201], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[202], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[203], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[204], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[205], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[206], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[207], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[208], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[209], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[210], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[211], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[212], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[213], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[214], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[215], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[216], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[217], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[218], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[219], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[220], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[221], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[222], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[223], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x21C):
                sheet.cell(row=row, column=1, value="ISPR7")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[224], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[225], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[226], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[227], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[228], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[229], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[230], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[231], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[232], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[233], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[234], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[235], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[236], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[237], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[238], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[239], offset=15, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x280):
                sheet.cell(row=row, column=1, value="ICPR0")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[0], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[1], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[2], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[3], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[4], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[5], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[6], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[7], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[8], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[9], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[10], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[11], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[12], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[13], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[14], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[15], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[16], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[17], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[18], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[19], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[20], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[21], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[22], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[23], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[24], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[25], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[26], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[27], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[28], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[29], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[30], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[31], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x284):
                sheet.cell(row=row, column=1, value="ICPR1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[32], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[33], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[34], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[35], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[36], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[37], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[38], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[39], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[40], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[41], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[42], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[43], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[44], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[45], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[46], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[47], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[48], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[49], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[50], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[51], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[52], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[53], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[54], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[55], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[56], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[57], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[58], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[59], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[60], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[61], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[62], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[63], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x288):
                sheet.cell(row=row, column=1, value="ICPR2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[64], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[65], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[66], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[67], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[68], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[69], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[70], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[71], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[72], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[73], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[74], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[75], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[76], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[77], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[78], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[79], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[80], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[81], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[82], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[83], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[84], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[85], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[86], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[87], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[88], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[89], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[90], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[91], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[92], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[93], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[94], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[95], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x28C):
                sheet.cell(row=row, column=1, value="ICPR3")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[96], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[97], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[98], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[99], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[100], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[101], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[102], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[103], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[104], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[105], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[106], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[107], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[108], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[109], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[110], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[111], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[112], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[113], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[114], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[115], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[116], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[117], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[118], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[119], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[120], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[121], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[122], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[123], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[124], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[125], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[126], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[127], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x290):
                sheet.cell(row=row, column=1, value="ICPR4")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[128], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[129], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[130], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[131], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[132], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[133], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[134], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[135], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[136], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[137], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[138], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[139], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[140], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[141], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[142], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[143], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[144], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[145], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[146], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[147], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[148], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[149], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[150], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[151], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[152], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[153], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[154], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[155], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[156], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[157], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[158], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[159], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x294):
                sheet.cell(row=row, column=1, value="ICPR5")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[160], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[161], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[162], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[163], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[164], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[165], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[166], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[167], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[168], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[169], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[170], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[171], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[172], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[173], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[174], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[175], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[176], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[177], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[178], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[179], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[180], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[181], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[182], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[183], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[184], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[185], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[186], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[187], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[188], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[189], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[190], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[191], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x298):
                sheet.cell(row=row, column=1, value="ICPR6")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[192], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[193], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[194], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[195], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[196], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[197], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[198], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[199], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[200], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[201], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[202], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[203], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[204], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[205], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[206], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[207], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[208], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[209], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[210], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[211], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[212], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[213], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[214], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[215], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[216], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[217], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[218], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[219], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[220], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[221], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[222], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[223], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x29C):
                sheet.cell(row=row, column=1, value="ICPR7")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[224], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[225], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[226], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[227], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[228], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[229], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[230], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[231], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[232], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[233], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[234], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[235], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[236], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[237], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[238], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[239], offset=15, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x300):
                sheet.cell(row=row, column=1, value="IABR0")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[0], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[1], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[2], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[3], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[4], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[5], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[6], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[7], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[8], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[9], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[10], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[11], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[12], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[13], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[14], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[15], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[16], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[17], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[18], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[19], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[20], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[21], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[22], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[23], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[24], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[25], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[26], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[27], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[28], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[29], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[30], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[31], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x304):
                sheet.cell(row=row, column=1, value="IABR1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[32], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[33], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[34], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[35], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[36], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[37], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[38], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[39], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[40], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[41], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[42], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[43], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[44], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[45], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[46], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[47], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[48], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[49], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[50], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[51], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[52], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[53], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[54], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[55], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[56], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[57], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[58], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[59], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[60], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[61], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[62], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[63], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x308):
                sheet.cell(row=row, column=1, value="IABR2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[64], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[65], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[66], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[67], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[68], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[69], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[70], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[71], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[72], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[73], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[74], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[75], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[76], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[77], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[78], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[79], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[80], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[81], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[82], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[83], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[84], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[85], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[86], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[87], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[88], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[89], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[90], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[91], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[92], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[93], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[94], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[95], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x30C):
                sheet.cell(row=row, column=1, value="IABR3")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[96], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[97], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[98], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[99], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[100], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[101], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[102], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[103], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[104], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[105], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[106], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[107], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[108], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[109], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[110], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[111], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[112], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[113], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[114], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[115], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[116], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[117], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[118], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[119], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[120], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[121], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[122], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[123], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[124], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[125], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[126], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[127], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x310):
                sheet.cell(row=row, column=1, value="IABR4")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[128], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[129], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[130], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[131], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[132], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[133], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[134], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[135], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[136], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[137], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[138], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[139], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[140], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[141], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[142], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[143], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[144], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[145], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[146], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[147], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[148], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[149], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[150], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[151], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[152], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[153], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[154], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[155], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[156], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[157], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[158], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[159], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x314):
                sheet.cell(row=row, column=1, value="IABR5")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[160], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[161], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[162], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[163], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[164], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[165], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[166], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[167], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[168], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[169], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[170], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[171], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[172], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[173], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[174], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[175], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[176], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[177], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[178], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[179], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[180], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[181], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[182], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[183], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[184], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[185], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[186], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[187], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[188], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[189], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[190], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[191], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x318):
                sheet.cell(row=row, column=1, value="IABR6")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[192], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[193], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[194], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[195], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[196], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[197], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[198], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[199], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[200], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[201], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[202], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[203], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[204], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[205], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[206], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[207], offset=15, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[208], offset=16, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[209], offset=17, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[210], offset=18, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[211], offset=19, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[212], offset=20, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[213], offset=21, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[214], offset=22, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[215], offset=23, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[216], offset=24, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[217], offset=25, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[218], offset=26, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[219], offset=27, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[220], offset=28, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[221], offset=29, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[222], offset=30, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[223], offset=31, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x31C):
                sheet.cell(row=row, column=1, value="IABR7")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name=self.NameList[224], offset=0, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[225], offset=1, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[226], offset=2, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[227], offset=3, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[228], offset=4, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[229], offset=5, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[230], offset=6, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[231], offset=7, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[232], offset=8, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[233], offset=9, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[234], offset=10, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[235], offset=11, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[236], offset=12, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[237], offset=13, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[238], offset=14, true_meaning="Enable", false_meaning="Disable"),
                    BitItem(name=self.NameList[239], offset=15, true_meaning="Enable", false_meaning="Disable")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x400):
                sheet.cell(row=row, column=1, value="IP0")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[0], offset=0, width=8),
                    IntItem(name=self.NameList[1], offset=8, width=8),
                    IntItem(name=self.NameList[2], offset=16, width=8),
                    IntItem(name=self.NameList[3], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x404):
                sheet.cell(row=row, column=1, value="IP1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[4], offset=0, width=8),
                    IntItem(name=self.NameList[5], offset=8, width=8),
                    IntItem(name=self.NameList[6], offset=16, width=8),
                    IntItem(name=self.NameList[7], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x408):
                sheet.cell(row=row, column=1, value="IP2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[8], offset=0, width=8),
                    IntItem(name=self.NameList[9], offset=8, width=8),
                    IntItem(name=self.NameList[10], offset=16, width=8),
                    IntItem(name=self.NameList[11], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x40C):
                sheet.cell(row=row, column=1, value="IP3")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[12], offset=0, width=8),
                    IntItem(name=self.NameList[13], offset=8, width=8),
                    IntItem(name=self.NameList[14], offset=16, width=8),
                    IntItem(name=self.NameList[15], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x410):
                sheet.cell(row=row, column=1, value="IP4")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[16], offset=0, width=8),
                    IntItem(name=self.NameList[17], offset=8, width=8),
                    IntItem(name=self.NameList[18], offset=16, width=8),
                    IntItem(name=self.NameList[19], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x414):
                sheet.cell(row=row, column=1, value="IP5")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[20], offset=0, width=8),
                    IntItem(name=self.NameList[21], offset=8, width=8),
                    IntItem(name=self.NameList[22], offset=16, width=8),
                    IntItem(name=self.NameList[23], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x418):
                sheet.cell(row=row, column=1, value="IP6")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[24], offset=0, width=8),
                    IntItem(name=self.NameList[25], offset=8, width=8),
                    IntItem(name=self.NameList[26], offset=16, width=8),
                    IntItem(name=self.NameList[27], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x41C):
                sheet.cell(row=row, column=1, value="IP7")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[28], offset=0, width=8),
                    IntItem(name=self.NameList[29], offset=8, width=8),
                    IntItem(name=self.NameList[30], offset=16, width=8),
                    IntItem(name=self.NameList[31], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x420):
                sheet.cell(row=row, column=1, value="IP8")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[32], offset=0, width=8),
                    IntItem(name=self.NameList[33], offset=8, width=8),
                    IntItem(name=self.NameList[34], offset=16, width=8),
                    IntItem(name=self.NameList[35], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x424):
                sheet.cell(row=row, column=1, value="IP9")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[36], offset=0, width=8),
                    IntItem(name=self.NameList[37], offset=8, width=8),
                    IntItem(name=self.NameList[38], offset=16, width=8),
                    IntItem(name=self.NameList[39], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x428):
                sheet.cell(row=row, column=1, value="IP10")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[40], offset=0, width=8),
                    IntItem(name=self.NameList[41], offset=8, width=8),
                    IntItem(name=self.NameList[42], offset=16, width=8),
                    IntItem(name=self.NameList[43], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x42C):
                sheet.cell(row=row, column=1, value="IP11")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[44], offset=0, width=8),
                    IntItem(name=self.NameList[45], offset=8, width=8),
                    IntItem(name=self.NameList[46], offset=16, width=8),
                    IntItem(name=self.NameList[47], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x430):
                sheet.cell(row=row, column=1, value="IP12")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[48], offset=0, width=8),
                    IntItem(name=self.NameList[49], offset=8, width=8),
                    IntItem(name=self.NameList[50], offset=16, width=8),
                    IntItem(name=self.NameList[51], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x434):
                sheet.cell(row=row, column=1, value="IP13")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[52], offset=0, width=8),
                    IntItem(name=self.NameList[53], offset=8, width=8),
                    IntItem(name=self.NameList[54], offset=16, width=8),
                    IntItem(name=self.NameList[55], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x438):
                sheet.cell(row=row, column=1, value="IP14")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[56], offset=0, width=8),
                    IntItem(name=self.NameList[57], offset=8, width=8),
                    IntItem(name=self.NameList[58], offset=16, width=8),
                    IntItem(name=self.NameList[59], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x43C):
                sheet.cell(row=row, column=1, value="IP15")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[60], offset=0, width=8),
                    IntItem(name=self.NameList[61], offset=8, width=8),
                    IntItem(name=self.NameList[62], offset=16, width=8),
                    IntItem(name=self.NameList[63], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x440):
                sheet.cell(row=row, column=1, value="IP16")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[64], offset=0, width=8),
                    IntItem(name=self.NameList[65], offset=8, width=8),
                    IntItem(name=self.NameList[66], offset=16, width=8),
                    IntItem(name=self.NameList[67], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x444):
                sheet.cell(row=row, column=1, value="IP17")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[68], offset=0, width=8),
                    IntItem(name=self.NameList[69], offset=8, width=8),
                    IntItem(name=self.NameList[70], offset=16, width=8),
                    IntItem(name=self.NameList[71], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x448):
                sheet.cell(row=row, column=1, value="IP18")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[72], offset=0, width=8),
                    IntItem(name=self.NameList[73], offset=8, width=8),
                    IntItem(name=self.NameList[74], offset=16, width=8),
                    IntItem(name=self.NameList[75], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x44C):
                sheet.cell(row=row, column=1, value="IP19")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[76], offset=0, width=8),
                    IntItem(name=self.NameList[77], offset=8, width=8),
                    IntItem(name=self.NameList[78], offset=16, width=8),
                    IntItem(name=self.NameList[79], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x450):
                sheet.cell(row=row, column=1, value="IP20")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[80], offset=0, width=8),
                    IntItem(name=self.NameList[81], offset=8, width=8),
                    IntItem(name=self.NameList[82], offset=16, width=8),
                    IntItem(name=self.NameList[83], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x454):
                sheet.cell(row=row, column=1, value="IP21")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[84], offset=0, width=8),
                    IntItem(name=self.NameList[85], offset=8, width=8),
                    IntItem(name=self.NameList[86], offset=16, width=8),
                    IntItem(name=self.NameList[87], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x458):
                sheet.cell(row=row, column=1, value="IP22")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[88], offset=0, width=8),
                    IntItem(name=self.NameList[89], offset=8, width=8),
                    IntItem(name=self.NameList[90], offset=16, width=8),
                    IntItem(name=self.NameList[91], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x45C):
                sheet.cell(row=row, column=1, value="IP23")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[92], offset=0, width=8),
                    IntItem(name=self.NameList[93], offset=8, width=8),
                    IntItem(name=self.NameList[94], offset=16, width=8),
                    IntItem(name=self.NameList[95], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x460):
                sheet.cell(row=row, column=1, value="IP24")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[96], offset=0, width=8),
                    IntItem(name=self.NameList[97], offset=8, width=8),
                    IntItem(name=self.NameList[98], offset=16, width=8),
                    IntItem(name=self.NameList[99], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x464):
                sheet.cell(row=row, column=1, value="IP25")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[100], offset=0, width=8),
                    IntItem(name=self.NameList[101], offset=8, width=8),
                    IntItem(name=self.NameList[102], offset=16, width=8),
                    IntItem(name=self.NameList[103], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x468):
                sheet.cell(row=row, column=1, value="IP26")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[104], offset=0, width=8),
                    IntItem(name=self.NameList[105], offset=8, width=8),
                    IntItem(name=self.NameList[106], offset=16, width=8),
                    IntItem(name=self.NameList[107], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x46C):
                sheet.cell(row=row, column=1, value="IP27")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[108], offset=0, width=8),
                    IntItem(name=self.NameList[109], offset=8, width=8),
                    IntItem(name=self.NameList[110], offset=16, width=8),
                    IntItem(name=self.NameList[111], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x470):
                sheet.cell(row=row, column=1, value="IP28")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[112], offset=0, width=8),
                    IntItem(name=self.NameList[113], offset=8, width=8),
                    IntItem(name=self.NameList[114], offset=16, width=8),
                    IntItem(name=self.NameList[115], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x474):
                sheet.cell(row=row, column=1, value="IP29")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[116], offset=0, width=8),
                    IntItem(name=self.NameList[117], offset=8, width=8),
                    IntItem(name=self.NameList[118], offset=16, width=8),
                    IntItem(name=self.NameList[119], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x478):
                sheet.cell(row=row, column=1, value="IP30")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[120], offset=0, width=8),
                    IntItem(name=self.NameList[121], offset=8, width=8),
                    IntItem(name=self.NameList[122], offset=16, width=8),
                    IntItem(name=self.NameList[123], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x47C):
                sheet.cell(row=row, column=1, value="IP31")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[124], offset=0, width=8),
                    IntItem(name=self.NameList[125], offset=8, width=8),
                    IntItem(name=self.NameList[126], offset=16, width=8),
                    IntItem(name=self.NameList[127], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x480):
                sheet.cell(row=row, column=1, value="IP32")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[128], offset=0, width=8),
                    IntItem(name=self.NameList[129], offset=8, width=8),
                    IntItem(name=self.NameList[130], offset=16, width=8),
                    IntItem(name=self.NameList[131], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x484):
                sheet.cell(row=row, column=1, value="IP33")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[132], offset=0, width=8),
                    IntItem(name=self.NameList[133], offset=8, width=8),
                    IntItem(name=self.NameList[134], offset=16, width=8),
                    IntItem(name=self.NameList[135], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x488):
                sheet.cell(row=row, column=1, value="IP34")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[136], offset=0, width=8),
                    IntItem(name=self.NameList[137], offset=8, width=8),
                    IntItem(name=self.NameList[138], offset=16, width=8),
                    IntItem(name=self.NameList[139], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x48C):
                sheet.cell(row=row, column=1, value="IP35")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[140], offset=0, width=8),
                    IntItem(name=self.NameList[141], offset=8, width=8),
                    IntItem(name=self.NameList[142], offset=16, width=8),
                    IntItem(name=self.NameList[143], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x490):
                sheet.cell(row=row, column=1, value="IP36")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[144], offset=0, width=8),
                    IntItem(name=self.NameList[145], offset=8, width=8),
                    IntItem(name=self.NameList[146], offset=16, width=8),
                    IntItem(name=self.NameList[147], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x494):
                sheet.cell(row=row, column=1, value="IP37")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[148], offset=0, width=8),
                    IntItem(name=self.NameList[149], offset=8, width=8),
                    IntItem(name=self.NameList[150], offset=16, width=8),
                    IntItem(name=self.NameList[151], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x498):
                sheet.cell(row=row, column=1, value="IP38")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[152], offset=0, width=8),
                    IntItem(name=self.NameList[153], offset=8, width=8),
                    IntItem(name=self.NameList[154], offset=16, width=8),
                    IntItem(name=self.NameList[155], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x49C):
                sheet.cell(row=row, column=1, value="IP39")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[156], offset=0, width=8),
                    IntItem(name=self.NameList[157], offset=8, width=8),
                    IntItem(name=self.NameList[158], offset=16, width=8),
                    IntItem(name=self.NameList[159], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4A0):
                sheet.cell(row=row, column=1, value="IP40")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[160], offset=0, width=8),
                    IntItem(name=self.NameList[161], offset=8, width=8),
                    IntItem(name=self.NameList[162], offset=16, width=8),
                    IntItem(name=self.NameList[163], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4A4):
                sheet.cell(row=row, column=1, value="IP41")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[164], offset=0, width=8),
                    IntItem(name=self.NameList[165], offset=8, width=8),
                    IntItem(name=self.NameList[166], offset=16, width=8),
                    IntItem(name=self.NameList[167], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4A8):
                sheet.cell(row=row, column=1, value="IP42")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[168], offset=0, width=8),
                    IntItem(name=self.NameList[169], offset=8, width=8),
                    IntItem(name=self.NameList[170], offset=16, width=8),
                    IntItem(name=self.NameList[171], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4AC):
                sheet.cell(row=row, column=1, value="IP43")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[172], offset=0, width=8),
                    IntItem(name=self.NameList[173], offset=8, width=8),
                    IntItem(name=self.NameList[174], offset=16, width=8),
                    IntItem(name=self.NameList[175], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4B0):
                sheet.cell(row=row, column=1, value="IP44")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[176], offset=0, width=8),
                    IntItem(name=self.NameList[177], offset=8, width=8),
                    IntItem(name=self.NameList[178], offset=16, width=8),
                    IntItem(name=self.NameList[179], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4B4):
                sheet.cell(row=row, column=1, value="IP45")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[180], offset=0, width=8),
                    IntItem(name=self.NameList[181], offset=8, width=8),
                    IntItem(name=self.NameList[182], offset=16, width=8),
                    IntItem(name=self.NameList[183], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4B8):
                sheet.cell(row=row, column=1, value="IP46")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[184], offset=0, width=8),
                    IntItem(name=self.NameList[185], offset=8, width=8),
                    IntItem(name=self.NameList[186], offset=16, width=8),
                    IntItem(name=self.NameList[187], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4BC):
                sheet.cell(row=row, column=1, value="IP47")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[188], offset=0, width=8),
                    IntItem(name=self.NameList[189], offset=8, width=8),
                    IntItem(name=self.NameList[190], offset=16, width=8),
                    IntItem(name=self.NameList[191], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4C0):
                sheet.cell(row=row, column=1, value="IP48")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[192], offset=0, width=8),
                    IntItem(name=self.NameList[193], offset=8, width=8),
                    IntItem(name=self.NameList[194], offset=16, width=8),
                    IntItem(name=self.NameList[195], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4C4):
                sheet.cell(row=row, column=1, value="IP49")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[196], offset=0, width=8),
                    IntItem(name=self.NameList[197], offset=8, width=8),
                    IntItem(name=self.NameList[198], offset=16, width=8),
                    IntItem(name=self.NameList[199], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4C8):
                sheet.cell(row=row, column=1, value="IP50")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[200], offset=0, width=8),
                    IntItem(name=self.NameList[201], offset=8, width=8),
                    IntItem(name=self.NameList[202], offset=16, width=8),
                    IntItem(name=self.NameList[203], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4CC):
                sheet.cell(row=row, column=1, value="IP51")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[204], offset=0, width=8),
                    IntItem(name=self.NameList[205], offset=8, width=8),
                    IntItem(name=self.NameList[206], offset=16, width=8),
                    IntItem(name=self.NameList[207], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4D0):
                sheet.cell(row=row, column=1, value="IP52")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[208], offset=0, width=8),
                    IntItem(name=self.NameList[209], offset=8, width=8),
                    IntItem(name=self.NameList[210], offset=16, width=8),
                    IntItem(name=self.NameList[211], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4D4):
                sheet.cell(row=row, column=1, value="IP53")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[212], offset=0, width=8),
                    IntItem(name=self.NameList[213], offset=8, width=8),
                    IntItem(name=self.NameList[214], offset=16, width=8),
                    IntItem(name=self.NameList[215], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4D8):
                sheet.cell(row=row, column=1, value="IP54")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[216], offset=0, width=8),
                    IntItem(name=self.NameList[217], offset=8, width=8),
                    IntItem(name=self.NameList[218], offset=16, width=8),
                    IntItem(name=self.NameList[219], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4DC):
                sheet.cell(row=row, column=1, value="IP55")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[220], offset=0, width=8),
                    IntItem(name=self.NameList[221], offset=8, width=8),
                    IntItem(name=self.NameList[222], offset=16, width=8),
                    IntItem(name=self.NameList[223], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4E0):
                sheet.cell(row=row, column=1, value="IP56")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[224], offset=0, width=8),
                    IntItem(name=self.NameList[225], offset=8, width=8),
                    IntItem(name=self.NameList[226], offset=16, width=8),
                    IntItem(name=self.NameList[227], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4E4):
                sheet.cell(row=row, column=1, value="IP57")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[228], offset=0, width=8),
                    IntItem(name=self.NameList[229], offset=8, width=8),
                    IntItem(name=self.NameList[230], offset=16, width=8),
                    IntItem(name=self.NameList[231], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4E8):
                sheet.cell(row=row, column=1, value="IP58")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[232], offset=0, width=8),
                    IntItem(name=self.NameList[233], offset=8, width=8),
                    IntItem(name=self.NameList[234], offset=16, width=8),
                    IntItem(name=self.NameList[235], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0x4EC):
                sheet.cell(row=row, column=1, value="IP59")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name=self.NameList[236], offset=0, width=8),
                    IntItem(name=self.NameList[237], offset=8, width=8),
                    IntItem(name=self.NameList[238], offset=16, width=8),
                    IntItem(name=self.NameList[239], offset=32, width=8)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0xD04):
                sheet.cell(row=row, column=1, value="ICSR")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                SysIntName = [
                    "Thread mode", "", "NMI", "HardFault",
                    "MemManage", "BusFault", "UsageFault", "",
                    "", "", "", "SVCall",
                    "DebugMonitor", "", "PendSV", "SysTick"]
                row = row + ProcessArray(array=[
                    IntItem(name="VECTACTIVE", offset=0, width=8, List=SysIntName + self.NameList),
                    BitItem(name="RETTOBASE", offset=11, true_meaning="no preempted exceptions", false_meaning="preempted active exceptions"),
                    IntItem(name="VECTPENDING", offset=12, width=8, List=SysIntName + self.NameList),
                    BitItem(name="ISRPENDING", offset=22, true_meaning="Interrupt pending", false_meaning="Interrupt not pending"),
                    BitItem(name="PENDSTSET", offset=26, true_meaning="SysTick pending", false_meaning="SysTick not pending"),
                    BitItem(name="PENDSVSET", offset=28, true_meaning="PendSV pending", false_meaning="PendSV not pending"),
                    BitItem(name="NMIPENDSET", offset=31, true_meaning="NMI pending", false_meaning="NMI not pending")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == scb_base + 0xD08):
                sheet.cell(row=row, column=1, value="VTOR")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
            else:
                row -= 1
