from typing import List, Tuple

import openpyxl

from BitItem import BitItem
from IModule import IModule
from IntItem import IntItem
from Parse import ProcessArray


class CAN(IModule):

    def __init__(self, name: str, address: int) -> None:
        self.Name = name
        self.Address = address

    def GetModuleName(self) -> str:
        return self.Name

    def __createArrary(self, blockId: int, trueMeaning: str, falseMeaning: str) -> List[BitItem]:
        array = []
        for i in range(blockId * 32, (blockId + 1) * 32):
            array.append(BitItem(name=f"MB{i}", offset=i, true_meaning=f"MB{i} " + trueMeaning, false_meaning=f"MB{i} " + falseMeaning))
        return array

    def UpdateSheet(self, data: List[Tuple[str, str]], sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        row = 1
        can_base = self.Address
        for item in data:
            row += 1
            addr = int(item[0], 16)
            value = int(item[1], 16)
            if (addr == can_base + 0x0000):
                sheet.cell(row=row, column=1, value="MCR")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Module Configuration")
                row = row + ProcessArray(array=[
                    BitItem(name="MDIS", offset=31, true_meaning="Module Disable", false_meaning="Module Enable"),
                    BitItem(name="FRZ", offset=30, true_meaning="Module Freeze", false_meaning="Module Running"),
                    BitItem(name="RFEN", offset=29, true_meaning="Legacy Rx FIFO Enable", false_meaning="Legacy Rx FIFO Disable"),
                    BitItem(name="HALT", offset=28, true_meaning="Halt Enable", false_meaning="No Request"),
                    BitItem(name="NOTRDY", offset=27, true_meaning="Not Ready", false_meaning="Running"),
                    BitItem(name="SOFTRST", offset=25, true_meaning="Soft Reset", false_meaning="No Reset"),
                    BitItem(name="FRZACK", offset=24, true_meaning="In Freeze Mode", false_meaning="Not In Freeze Mode"),
                    BitItem(name="SUPV", offset=23, true_meaning="Supervisor Only", false_meaning="All Mode"),
                    BitItem(name="WRNEN", offset=21, true_meaning="Warning Interrupt Enable", false_meaning="Warning Interrupt Disable"),
                    BitItem(name="LPMACK", offset=20, true_meaning="In Low-Power Mode", false_meaning="Not In Low-Power Mode"),
                    BitItem(name="SRXDIS", offset=17, true_meaning="Self-Reception Disable", false_meaning="Self-Reception Enable"),
                    BitItem(name="IRMQ", offset=16, true_meaning="Individual RX Masking and Queue Enable", false_meaning="Individual RX Masking and Queue Disable"),
                    BitItem(name="DMA", offset=15, true_meaning="DMA Enable", false_meaning="DMA Disable"),
                    BitItem(name="LPRIOEN", offset=13, true_meaning="Local Priority Enable", false_meaning="Local Priority Disable"),
                    BitItem(name="AEN", offset=12, true_meaning="Abort Enable", false_meaning="Abort Disable"),
                    BitItem(name="FDEN", offset=11, true_meaning="CANFD Enable", false_meaning="CANFD Disable"),
                    BitItem(name="TPOV", offset=10, true_meaning="TX Pin Set 1", false_meaning="TX Pin Set 0"),
                    BitItem(name="TPOE", offset=7, true_meaning="TX Pin Force Enable", false_meaning="TX Pin Force Disable"),
                    IntItem(name="IDAM", offset=8, width=2, enumStr=["Format A", "Format B", "Format C", "Format D"]),
                    IntItem(name="MAXMB", offset=0, width=7, description="Number of the Last Message Buffer"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0004):
                sheet.cell(row=row, column=1, value="CTRL1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Control 1")
                row = row + ProcessArray(array=[
                    IntItem(name="PRESDIV", offset=24, width=8, description="Prescaler Division Factor"),
                    IntItem(name="RJW", offset=22, width=2, description="Resync Jump Width"),
                    IntItem(name="PSEG1", offset=19, width=3, description="Phase Segment 1"),
                    IntItem(name="PSEG2", offset=16, width=3, description="Phase Segment 2"),
                    BitItem(name="BOFFMSK", offset=15, true_meaning="BusOff Interrupt Enable", false_meaning="BusOff Interrupt Disable"),
                    BitItem(name="ERRMSK", offset=14, true_meaning="Error Interrupt Enable", false_meaning="Error Interrupt Disable"),
                    BitItem(name="LPB", offset=12, true_meaning="Loopback Enable", false_meaning="Loopback Disable"),
                    BitItem(name="TWRNMSK", offset=11, true_meaning="TX Warning Interrupt Enable", false_meaning="TX Warning Interrupt Disable"),
                    BitItem(name="RWRNMSK", offset=10, true_meaning="RX Warning Interrupt Enable", false_meaning="RX Warning Interrupt Disable"),
                    BitItem(name="ROM", offset=8, true_meaning="Restricted Mode Enable", false_meaning="Restricted Mode Disable"),
                    BitItem(name="SMP", offset=7, true_meaning="3 Bits Sample", false_meaning="1 Bit Sample"),
                    BitItem(name="BOFFREC", offset=6, true_meaning="BusOff Recovery Enable", false_meaning="BusOff Recovery Disable"),
                    BitItem(name="TSYN", offset=5, true_meaning="Time Sync Enable", false_meaning="Time Sync Disable"),
                    BitItem(name="LBUF", offset=4, true_meaning="Low Buffer First", false_meaning="High Buffer First"),
                    BitItem(name="LOM", offset=3, true_meaning="Listen Only Mode", false_meaning="Normal Mode"),
                    IntItem(name="PROPSEG", offset=0, width=3, description="Propagation Segment"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0008):
                sheet.cell(row=row, column=1, value="TIMER")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Free-Running Timer")
                row = row + ProcessArray(array=[
                    IntItem(name="TIMER", offset=0, width=16, description="Timer Value"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0010):
                sheet.cell(row=row, column=1, value="RXMGMASK")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="RX Message Buffers Global Mask")
            elif (addr == can_base + 0x0014):
                sheet.cell(row=row, column=1, value="RX14MASK")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Receive 14 Mask")
            elif (addr == can_base + 0x0018):
                sheet.cell(row=row, column=1, value="RX15MASK")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Receive 15 Mask")
            elif (addr == can_base + 0x001C):
                sheet.cell(row=row, column=1, value="ECR")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Error Counter")
                row = row + ProcessArray(array=[
                    IntItem(name="RXERRCNT_FAST", offset=24, width=8, description="Receive Error Counter for Fast Bits"),
                    IntItem(name="TXERRCNT_FAST", offset=16, width=8, description="Transmit Error Counter for Fast Bits"),
                    IntItem(name="RXERRCNT", offset=8, width=8, description="Receive Error Counter for Standard Bits"),
                    IntItem(name="TXERRCNT", offset=0, width=8, description="Transmit Error Counter for Standard Bits"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0020):
                sheet.cell(row=row, column=1, value="ESR1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Error Status 1")
                row = row + ProcessArray(array=[
                    BitItem(name="BIT1ERR_FAST", offset=31, true_meaning="Fast Bit 1 Error", false_meaning=""),
                    BitItem(name="BIT0ERR_FAST", offset=30, true_meaning="Fast Bit 0 Error", false_meaning=""),
                    BitItem(name="CRCERR_FAST", offset=28, true_meaning="Fast CRC Error", false_meaning=""),
                    BitItem(name="FRMERR_FAST", offset=27, true_meaning="Fast Form Error", false_meaning=""),
                    BitItem(name="STFERR_FAST", offset=26, true_meaning="Fast Stuffing Error", false_meaning=""),
                    BitItem(name="PTA", offset=23, true_meaning="Passive to Active State", false_meaning=""),
                    BitItem(name="ATP", offset=22, true_meaning="Active to Passive State", false_meaning=""),
                    BitItem(name="ERROVR", offset=21, true_meaning="Overrun", false_meaning=""),
                    BitItem(name="ERRINT_FAST", offset=20, true_meaning="Fast Error Interrupt", false_meaning=""),
                    BitItem(name="BOFFDONEINT", offset=19, true_meaning="BusOff Done", false_meaning=""),
                    BitItem(name="SYNCH", offset=18, true_meaning="CAN Synchronized", false_meaning=""),
                    BitItem(name="TWRNINT", offset=17, true_meaning="TX Warning Interrupt", false_meaning=""),
                    BitItem(name="RWRNINT", offset=16, true_meaning="RX Warning Interrupt", false_meaning=""),
                    BitItem(name="BIT1ERR", offset=15, true_meaning="Bit 1 Error", false_meaning=""),
                    BitItem(name="BIT0ERR", offset=14, true_meaning="Bit 0 Error", false_meaning=""),
                    BitItem(name="ACKERR", offset=13, true_meaning="Acknowledge Error", false_meaning=""),
                    BitItem(name="CRCERR", offset=12, true_meaning="CRC Error", false_meaning=""),
                    BitItem(name="FRMERR", offset=11, true_meaning="Form Error", false_meaning=""),
                    BitItem(name="STFERR", offset=10, true_meaning="Stuffing Error", false_meaning=""),
                    BitItem(name="TXWRN", offset=9, true_meaning="TX Warning", false_meaning=""),
                    BitItem(name="RXWRN", offset=8, true_meaning="RX Warning", false_meaning=""),
                    BitItem(name="IDLE", offset=7, true_meaning="IDLE", false_meaning="Not IDLE"),
                    BitItem(name="TX", offset=6, true_meaning="Transmitting", false_meaning="Not Transmitting"),
                    IntItem(name="FLTCONF", offset=4, width=2, enumStr=["Error Active", "Error Passive", "Bus Off", "Bus Off"]),
                    BitItem(name="RX", offset=3, true_meaning="Receiving", false_meaning="Not Receiving"),
                    BitItem(name="BOFFINT", offset=2, true_meaning="BusOff Interrupt", false_meaning=""),
                    BitItem(name="ERRINT", offset=1, true_meaning="Error Interrupt", false_meaning=""),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0024):
                sheet.cell(row=row, column=1, value="IMASK2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Interrupt Mask 2")
                row = row + ProcessArray(array=self.__createArrary(blockId=1, trueMeaning="Interrupt Enable", falseMeaning="Interrupt Disable"),
                    value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0028):
                sheet.cell(row=row, column=1, value="IMASK1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Interrupt Mask 1")
                row = row + ProcessArray(array=self.__createArrary(blockId=0, trueMeaning="Interrupt Enable", falseMeaning="Interrupt Disable"),
                    value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x002C):
                sheet.cell(row=row, column=1, value="IFLAG2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Interrupt Flag 2")
                row = row + ProcessArray(array=self.__createArrary(blockId=1, trueMeaning="Interrupt Exist", falseMeaning="Interrupt Not Exist"),
                    value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0030):
                sheet.cell(row=row, column=1, value="IFLAG1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Interrupt Flag 1")
                row = row + ProcessArray(array=self.__createArrary(blockId=0, trueMeaning="Interrupt Exist", falseMeaning="Interrupt Not Exist"),
                    value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0034):
                sheet.cell(row=row, column=1, value="CTRL2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Control 2")
                row = row + ProcessArray(array=[
                    BitItem(name="ERRMSK_FAST", offset=31, true_meaning="Error Interrupt Enable", false_meaning="Error Interrupt Disable"),
                    BitItem(name="BOFFDONEMSK", offset=30, true_meaning="BusOff Done Interrupt Enable", false_meaning="BusOff Done Interrupt Disable"),
                    BitItem(name="ECRWRE", offset=29, true_meaning="MECR Write Enable", false_meaning="MECR Write Disable"),
                    BitItem(name="WRMFRZ", offset=28, true_meaning="Write Mem in Freeze Enable", false_meaning="Write Mem in Freeze Disable"),
                    IntItem(name="RFFN", offset=24, width=4, description="Number of Legacy Receive FIFO Filters"),
                    IntItem(name="TASD", offset=19, width=5, description="Transmission Arbitration Start Delay"),
                    BitItem(name="MRP", offset=18, true_meaning="MB Priority High", false_meaning="MB Priority Low"),
                    BitItem(name="RRS", offset=17, true_meaning="Remote Request Stored", false_meaning="Remote Request Generated"),
                    BitItem(name="EACEN", offset=16, true_meaning="EAC Enable", false_meaning="EAC Disable"),
                    BitItem(name="TIMER_SRC", offset=15, true_meaning="External", false_meaning="CAN Bit"),
                    BitItem(name="PREXCEN", offset=14, true_meaning="Protocol Exception Enable", false_meaning="Protocol Exception Disable"),
                    BitItem(name="BTE", offset=13, true_meaning="Bit Timing Expansion Enable", false_meaning="Bit Timing Expansion Disable"),
                    BitItem(name="ISOCANFDEN", offset=12, true_meaning="ISO CAN FD Enable", false_meaning="ISO CAN FD Disable"),
                    BitItem(name="EDFLTDIS", offset=11, true_meaning="Edge Filter Disable", false_meaning="Edge Filter Enable"),
                    BitItem(name="FLT_RXN", offset=10, true_meaning="Fault Reaction Enable", false_meaning="Fault Reaction Disable"),
                    IntItem(name="MBTSBASE", offset=8, width=2, enumStr=["TIMER", "Lower 16bits Timer", "Upper 16bits Timer", "Reserved"]),
                    IntItem(name="TSTAMPCAP", offset=6, width=2, enumStr=["Disabled", "End of Frame", "Start of Frame", "Start CAN / Res CANFD"]),
                    IntItem(name="RETRY", offset=2, width=3, enumStr=["No Retry", "1", "2", "3", "4", "5", "6", "Unlimited"]),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0038):
                sheet.cell(row=row, column=1, value="ESR2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Error Status 2")
                row = row + ProcessArray(array=[
                    IntItem(name="LPTM", offset=16, width=7, description="Lowest Priority TX Message Buffer"),
                    BitItem(name="VPS", offset=14, true_meaning="Priority Valid", false_meaning="Priority Invalid"),
                    BitItem(name="IMB", offset=13, true_meaning="LPTM not inactive", false_meaning="MB inactive"),
                    BitItem(name="RX_PIN_ST", offset=12, true_meaning="RX 1", false_meaning="RX 0"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0044):
                sheet.cell(row=row, column=1, value="CRCR")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Cyclic Redundancy Check")
                row = row + ProcessArray(array=[
                    IntItem(name="MBCRC", offset=16, width=7, description="CRC Message Buffer"),
                    IntItem(name="TXCRC", offset=0, width=15, description="Transmitted CRC value"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0050):
                sheet.cell(row=row, column=1, value="CBT")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="CAN Bit Timing")
                row = row + ProcessArray(array=[
                    BitItem(name="BTF", offset=31, true_meaning="Bit Timing Format Enable", false_meaning="Bit Timing Format Disable"),
                    IntItem(name="EPRESDIV", offset=21, width=10, description="Extended Prescaler Division Factor"),
                    IntItem(name="ERJW", offset=16, width=5, description="Extended Resync Jump Width"),
                    IntItem(name="EPROPSEG", offset=10, width=6, description="Extended Propagation Segment"),
                    IntItem(name="EPSEG1", offset=5, width=5, description="Extended Phase Segment 1"),
                    IntItem(name="EPSEG2", offset=0, width=5, description="Extended Phase Segment 2"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x006C):
                sheet.cell(row=row, column=1, value="IMASK3")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Interrupt Mask 3")
                row = row + ProcessArray(array=self.__createArrary(blockId=2, trueMeaning="Interrupt Enable", falseMeaning="Interrupt Disable"),
                    value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0074):
                sheet.cell(row=row, column=1, value="IFLAG3")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Interrupt Flag 3")
                row = row + ProcessArray(array=self.__createArrary(blockId=2, trueMeaning="Interrupt Exist", falseMeaning="Interrupt Not Exist"),
                    value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr >= can_base + 0x0880 and addr <= can_base + 0x09FC):
                sheet.cell(row=row, column=1, value=f"RXIMR{(addr - can_base - 0x0880) // 4}")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Receive Individual Mask")
            elif (addr == can_base + 0x0BF0):
                sheet.cell(row=row, column=1, value="EPRS")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Enhanced CAN Bit Timing Prescalers")
                row = row + ProcessArray(array=[
                    IntItem(name="EDPRESDIV", offset=16, width=10, description="Extended Data Phase Prescaler Division Factor"),
                    IntItem(name="ENPRESDIV", offset=0, width=10, description="Extended Nominal Prescaler Division Factor"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0BF4):
                sheet.cell(row=row, column=1, value="ENCBT")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Enhanced Nominal CAN Bit Timing")
                row = row + ProcessArray(array=[
                    IntItem(name="NRJW", offset=22, width=7, description="Nominal Resynchronization Jump Width"),
                    IntItem(name="NTSEG2", offset=12, width=7, description="Nominal Time Segment 2"),
                    IntItem(name="NTSEG1", offset=0, width=8, description="Nominal Time Segment 1"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0BF8):
                sheet.cell(row=row, column=1, value="EDCBT")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Enhanced Data Phase CAN Bit Timing")
                row = row + ProcessArray(array=[
                    IntItem(name="DRJW", offset=22, width=4, description="Data Phase Resynchronization Jump Width"),
                    IntItem(name="DTSEG2", offset=12, width=4, description="Data Phase Segment 2"),
                    IntItem(name="DTSEG1", offset=0, width=5, description="Data Phase Segment 1"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0BFC):
                sheet.cell(row=row, column=1, value="ETDC")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Enhanced TransDelay Compensation")
                row = row + ProcessArray(array=[
                    BitItem(name="ETDCEN", offset=31, true_meaning="TransDelay Compensation Enable", false_meaning="TransDelay Compensation Disable"),
                    BitItem(name="TDMDIS", offset=30, true_meaning="TransDelay Measurement Disable", false_meaning="TransDelay Measurement Enable"),
                    IntItem(name="ETDCOFF", offset=16, width=7, description="Enhanced TransDelay Compensation Offset"),
                    BitItem(name="ETDCFAIL", offset=15, true_meaning="TransDelay Compensation Fail", false_meaning="TransDelay Compensation Success"),
                    IntItem(name="ETDCVAL", offset=0, width=8, description="Enhanced TransDelay Compensation Value"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0C00):
                sheet.cell(row=row, column=1, value="FDCTRL")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="CAN FD Control")
                row = row + ProcessArray(array=[
                    BitItem(name="FDRATE", offset=31, true_meaning="Bit Rate Switch Enable", false_meaning="Bit Rate Switch Disable"),
                    IntItem(name="MBDSR2", offset=22, width=2, enumStr=["8B", "16B", "32B", "64B"]),
                    IntItem(name="MBDSR1", offset=19, width=2, enumStr=["8B", "16B", "32B", "64B"]),
                    IntItem(name="MBDSR0", offset=16, width=2, enumStr=["8B", "16B", "32B", "64B"]),
                    BitItem(name="TDCEN", offset=15, true_meaning="TransDelay Compensation Enable", false_meaning="TransDelay Compensation Disable"),
                    BitItem(name="TDCFAIL", offset=14, true_meaning="TransDelay Compensation Fail", false_meaning="TransDelay Compensation Success"),
                    IntItem(name="TDCOFF", offset=8, width=5, description="TransDelay Compensation Offset"),
                    IntItem(name="TDCVAL", offset=0, width=6, description="TransDelay Compensation Value"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0C04):
                sheet.cell(row=row, column=1, value="FDCBT")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="CAN FD Bit Timing")
                row = row + ProcessArray(array=[
                    IntItem(name="FPRESDIV", offset=20, width=10, description="Fast Prescaler Division Factor"),
                    IntItem(name="FRJW", offset=16, width=3, description="Fast Resync Jump Width"),
                    IntItem(name="FPROPSEG", offset=10, width=5, description="Fast Propagation Segment"),
                    IntItem(name="FPSEG1", offset=5, width=3, description="Fast Phase Segment 1"),
                    IntItem(name="FPSEG2", offset=0, width=3, description="Fast Phase Segment 2"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0C08):
                sheet.cell(row=row, column=1, value="FDCRC")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="CAN FD CRC")
                row = row + ProcessArray(array=[
                    IntItem(name="FD_MBCRC", offset=24, width=7, description="CRC Message Buffer Number for FD_TXCRC"),
                    IntItem(name="FD_TXCRC", offset=0, width=21, description="Extended Transmitted CRC value"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0C0C):
                sheet.cell(row=row, column=1, value="ERFCR")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Enhanced RX FIFO Control")
                row = row + ProcessArray(array=[
                    BitItem(name="ERFEN", offset=31, true_meaning="Enhanced RX FIFO Enable", false_meaning="Enhanced RX FIFO Disable"),
                    IntItem(name="DMALW", offset=26, width=5, description="Enhanced RX FIFO Control"),
                    IntItem(name="NEXIF", offset=16, width=7, description="Number of Extended ID Filter Elements"),
                    IntItem(name="NFE", offset=8, width=6, description="Number of Enhanced RX FIFO Filter Elements"),
                    IntItem(name="ERFWM", offset=0, width=5, description="Enhanced RX FIFO Watermark"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0C10):
                sheet.cell(row=row, column=1, value="ERFIER")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Enhanced RX FIFO Interrupt Enable")
                row = row + ProcessArray(array=[
                    BitItem(name="ERFUFWIE", offset=31, true_meaning="ERXFIFO Underflow Interrupt Enable", false_meaning="ERXFIFO Underflow Interrupt Disable"),
                    BitItem(name="ERFOVFIE", offset=30, true_meaning="ERXFIFO Overflow Interrupt Enable", false_meaning="ERXFIFO Overflow Interrupt Disable"),
                    BitItem(name="ERFWMIIE", offset=29, true_meaning="ERXFIFO Watermark Indication Interrupt Enable", false_meaning="ERXFIFO Watermark Indication Interrupt Disable"),
                    BitItem(name="ERFDAIE", offset=28, true_meaning="ERXFIFO Data Available Interrupt Enable", false_meaning="ERXFIFO Data Available Interrupt Disable"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == can_base + 0x0C14):
                sheet.cell(row=row, column=1, value="ERFSR")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="Enhanced RX FIFO Status")
                row = row + ProcessArray(array=[
                    BitItem(name="ERFUFW", offset=31, true_meaning="ERXFIFO Underflow Flag", false_meaning=""),
                    BitItem(name="ERFOVF", offset=30, true_meaning="ERXFIFO Overflow Flag", false_meaning=""),
                    BitItem(name="ERFWMI", offset=29, true_meaning="ERXFIFO Watermark Indication Flag", false_meaning=""),
                    BitItem(name="ERFDA", offset=28, true_meaning="ERXFIFO Data Available Flag", false_meaning=""),
                    BitItem(name="ERFE", offset=17, true_meaning="ERXFIFO Empty Flag", false_meaning=""),
                    BitItem(name="ERFF", offset=16, true_meaning="ERXFIFO Full Flag", false_meaning=""),
                    IntItem(name="ERFEL", offset=0, width=6, description="Enhanced RX FIFO Elements"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            else:
                row -= 1
