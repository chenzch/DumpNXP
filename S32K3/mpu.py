from typing import List, Tuple

import openpyxl

from BitItem import BitItem
from IModule import IModule
from IntItem import IntItem
from Parse import ProcessArray


class MPU(IModule):

    def __init__(self, name: str, address: int) -> None:
        self.Name = name
        self.Address = address

    def GetModuleName(self) -> str:
        return self.Name

    def UpdateSheet(self, data: List[Tuple[str, str]], sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        row = 1
        mpu_base = self.Address
        for item in data:
            row += 1
            addr = int(item[0], 16)
            value = int(item[1], 16)
            if (addr == mpu_base + 0xD94):
                sheet.cell(row=row, column=1, value="CTRL")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name="PRIVDEFENA",offset=2, true_meaning="Enable default memory map", false_meaning="Disable default memory map"),
                    BitItem(name="HFNMIENA", offset=1, true_meaning="MPU Enabled in hard fault", false_meaning="MPU Disabled in hard fault"),
                    BitItem(name="ENABLE", offset=0, true_meaning="MPU Enabled", false_meaning="MPU Disabled"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr >= mpu_base + 0xD9C and addr <= mpu_base + 0xE18):
                regionIndex = (addr - mpu_base - 0xD9C) // 8
                if (addr & 0x8) == 1:
                    sheet.cell(row=row, column=1, value=f"MPU_RBAR{regionIndex}")
                    sheet.cell(row=row, column=2, value=f"{value:08X}")
                    row = row + ProcessArray(array=[
                        IntItem(name="PLL_ODIV2", offset=25, width=6),
                        IntItem(name="PLL_RDIV", offset=12, width=3),
                        IntItem(name="PLL_MFI", offset=0, width=8),
                    ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
                else:
                    sheet.cell(row=row, column=1, value=f"MPU_RASR{regionIndex}")
                    sheet.cell(row=row, column=2, value=f"{value:08X}")
                    row = row + ProcessArray(array=[
                        IntItem(name="PLL_ODIV2", offset=25, width=6),
                        IntItem(name="PLL_RDIV", offset=12, width=3),
                        IntItem(name="PLL_MFI", offset=0, width=8),
                    ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            else:
                row -= 1
# https://developer.arm.com/documentation/dui0646/c/Cortex-M7-Peripherals/Optional-Memory-Protection-Unit/MPU-Region-Base-Address-Register?lang=en
