from typing import List, Tuple

import openpyxl

from BitItem import BitItem
from IModule import IModule
from IntItem import IntItem
from Parse import ProcessArray


class PLL(IModule):

    def __init__(self, name: str, address: int) -> None:
        self.Name = name
        self.Address = address

    def GetModuleName(self) -> str:
        return self.Name

    def UpdateSheet(self, data: List[Tuple[str, str]], sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        row = 1
        pll_base = self.Address
        for item in data:
            row += 1
            addr = int(item[0], 16)
            value = int(item[1], 16)
            if (addr == pll_base + 0x04):
                sheet.cell(row=row, column=1, value="PLL Status")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name="LOCK", offset=2, true_meaning="PLL locked", false_meaning="PLL unlocked"),
                    BitItem(name="LOL",offset=3, true_meaning="Loss of lock detected", false_meaning="")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == pll_base + 0x08):
                sheet.cell(row=row, column=1, value="PLL Divider")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name="PLL_MFI", offset=0, width=8),
                    IntItem(name="PLL_RDIV", offset=12, width=3),
                    IntItem(name="PLL_ODIV2", offset=25, width=6)
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            else:
                row -= 1
