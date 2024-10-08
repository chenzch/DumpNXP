from typing import List, Tuple

import openpyxl

from BitItem import BitItem
from IModule import IModule
from IntItem import IntItem
from Parse import ProcessArray


class MC_CGM(IModule):
    def __init__(self, name: str, address: int) -> None:
        self.Name = name
        self.Address = address

    def GetModuleName(self) -> str:
        return self.Name

    def UpdateSheet(self, data: List[Tuple[str, str]], sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        row = 1
        mc_cgm_base = self.Address
        for item in data:
            row += 1
            addr = int(item[0], 16)
            value = int(item[1], 16)
            if (addr == mc_cgm_base + 0x304):
                sheet.cell(row=row, column=1, value="MUX_0_CSS")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name="MUX_0_SEL", offset=24, width=4),
                    IntItem(name="MUL_0_SWTRG", offset=17, width=3),
                    BitItem(name="SWIP", offset=16, true_meaning="Clock switching", false_meaning="Clock switched"),
                    BitItem(name="SAFE_SW", offset=3, true_meaning="Safe clock switch requested", false_meaning=""),
                    BitItem(name="CLK_SW", offset=2, true_meaning="Clock switch requested", false_meaning=""),
                    BitItem(name="RAMPDOWN", offset=1, true_meaning="Ramp-down requested", false_meaning=""),
                    BitItem(name="RAMPUP", offset=0, true_meaning="Ramp-up requested", false_meaning=""),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == mc_cgm_base + 0x308):
                sheet.cell(row=row, column=1, value="MUX_0_DC0")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name="DIV", offset=16, width=3),
                    BitItem(name="DE", offset=31, true_meaning="Divider is enabled", false_meaning="Divider is disabled"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == mc_cgm_base + 0x30C):
                sheet.cell(row=row, column=1, value="MUX_0_DC1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name="DIV", offset=16, width=3),
                    BitItem(name="DE", offset=31, true_meaning="Divider is enabled", false_meaning="Divider is disabled"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == mc_cgm_base + 0x310):
                sheet.cell(row=row, column=1, value="MUX_0_DC2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name="DIV", offset=16, width=3),
                    BitItem(name="DE", offset=31, true_meaning="Divider is enabled", false_meaning="Divider is disabled"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == mc_cgm_base + 0x314):
                sheet.cell(row=row, column=1, value="MUX_0_DC3")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name="DIV", offset=16, width=3),
                    BitItem(name="DE", offset=31, true_meaning="Divider is enabled", false_meaning="Divider is disabled"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == mc_cgm_base + 0x318):
                sheet.cell(row=row, column=1, value="MUX_0_DC4")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    IntItem(name="DIV", offset=16, width=3),
                    BitItem(name="DE", offset=31, true_meaning="Divider is enabled", false_meaning="Divider is disabled"),
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            else:
                row -= 1
