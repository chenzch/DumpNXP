from typing import List, Tuple

import openpyxl

from BitItem import BitItem
from IModule import IModule
from Parse import ProcessArray


class Clock(IModule):
    def __init__(self, name: str, address: int) -> None:
        self.Name = name
        self.Address = address

    def GetModuleName(self) -> str:
        return self.Name

    def UpdateSheet(self, data: List[Tuple[str, str]], sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        row = 1
        sirc_base = self.Address
        sxosc_base = sirc_base + 0x00004000
        firc_base = sxosc_base + 0x00004000
        fxosc_base = firc_base + 0x00004000
        for item in data:
            row += 1
            addr = int(item[0], 16)
            value = int(item[1], 16)
            if (addr == sirc_base + 0x04):
                sheet.cell(row=row, column=1, value="SIRC")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name="STATUS", offset=0, true_meaning="ON", false_meaning="OFF")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == sirc_base + 0x0C):
                sheet.cell(row=row, column=1, value="SIRC enabled in standby")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name="STDBY_EN", offset=8, true_meaning="SIRC enabled in standby", false_meaning="SIRC disabled in standby")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == firc_base + 0x04):
                sheet.cell(row=row, column=1, value="FIRC")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name="STATUS", offset=0, true_meaning="ON", false_meaning="OFF")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == firc_base + 0x08):
                sheet.cell(row=row, column=1, value="FIRC enabled in standby")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name="STDBY_EN", offset=0, true_meaning="FIRC enabled in standby", false_meaning="FIRC disabled in standby")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == sxosc_base + 0x04):
                sheet.cell(row=row, column=1, value="SXOSC")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name="OSC_STAT", offset=31, true_meaning="ON", false_meaning="OFF")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == fxosc_base + 0x04):
                sheet.cell(row=row, column=1, value="FXOSC")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name="OSC_STAT", offset=31, true_meaning="ON", false_meaning="OFF")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            else:
                row -= 1
