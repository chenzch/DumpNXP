from typing import List, Tuple

import openpyxl

from IModule import IModule
from IParseItem import IParseItem
from BitItem import BitItem
from Parse import ProcessArray

class MC_ME(IModule):
    def __init__(self, name: str, address: int) -> None:
        self.Name = name
        self.Address = address

    def GetModuleName(self) -> str:
        return self.Name

    def __createArray(self, data: List[str]) -> List[IParseItem]:
        array = []
        index = 31
        for item in data:
            array.append(BitItem(name=item, offset=index, true_meaning=item + " clock is active", false_meaning=item + " clock is inactive"))
            index -= 1
        return array

    def UpdateSheet(self, data: List[Tuple[str, str]], sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        row = 1
        mcme_base = self.Address
        for item in data:
            row += 1
            addr = int(item[0], 16)
            value = int(item[1], 16)
            if (addr == mcme_base + 0x108):
                sheet.cell(row=row, column=1, value="MCME_PRTN0_STAT")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name="PCS", offset=0, true_meaning="Partition 0 clock is active", false_meaning="Partition 0 clock is inactive")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == mcme_base + 0x308):
                sheet.cell(row=row, column=1, value="MCME_PRTN1_STAT")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=[
                    BitItem(name="PCS", offset=0, true_meaning="Partition 1 clock is active", false_meaning="Partition 1 clock is inactive")
                ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == mcme_base + 0x114):
                sheet.cell(row=row, column=1, value="MCME_PRTN0_COFB1_STAT")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=self.__createArray(data=[
                    "Block63", "Block62", "Block61", "Block60", "Block59", "Block58", "Block57", "Block56", "Block55", "Block54", "Block53", "Block52",
                    "MU_4", "MU_4", "MU_3", "Block48", "MU_2", "MU_2", "PIT_1", "PIT_0", "Block43", "ADC_2", "ADC_1", "ADC_0", "LCU_1", "LCU_0",
                    "Block37", "eMIOS_2", "eMIOS_1", "eMIOS_0", "BCTU", "TRGMUX"]), value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr >= mcme_base + 0x310 and addr <= mcme_base + 0x31C):
                blockIndex = (addr - mcme_base - 0x310) // 4
                sheet.cell(row=row, column=1, value=["MCME_PRTN1_COFB0_STAT", "MCME_PRTN1_COFB1_STAT", "MCME_PRTN1_COFB2_STAT", "MCME_PRTN1_COFB3_STAT"][blockIndex])
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                row = row + ProcessArray(array=self.__createArray(data=[
                    ["INTM", "XRDC", "STM0", "SWT0", "PFC_alt", "PFC", "PRAM0", "MSCM", "ERM_0", "Block22", "SDA-AP", "Block20",
                    "Block19", "Block18", "Block17", "Block16", "eDMA_TCD_11", "eDMA_TCD_10", "eDMA_TCD_9", "eDMA_TCD_8", "eDMA_TCD_7",
                    "eDMA_TCD_6", "eDMA_TCD_5", "eDMA_TCD_4", "eDMA_TCD_3", "eDMA_TCD_2", "eDMA_TCD_1", "eDMA_TCD_0", "eDMA_CONTROL",
                    "Periph_XBIC", "System_XBIC", "AXBS"],
                    ["PIT_2", "SIUL_VIRTWRAPPER_PDAC4", "SIUL_VIRTWRAPPER_PDAC4", "FMU_alt", "FMU", "PMC", "PLL_2", "PLL", "MC_ME", "MC_CGM", "FXOSC", "FIRC",
                    "SXOSC", "SIRC", "TSPC", "Block48", "CMU_0-5", "Block46", "WKPU", "Block44", "DCM", "SIUL_VIRTWRAPPER_PDAC3", "SIUL_VIRTWRAPPER_PDAC2_M7_1",
                    "SIUL_VIRTWRAPPER_PDAC2_M7_1", "SIUL_VIRTWRAPPER_PDAC1_M7_0", "SIUL_VIRTWRAPPER_PDAC1_M7_0", "SIUL_VIRTWRAPPER_PDAC0_HSE", "SIUL_VIRTWRAPPER_PDAC0_HSE",
                    "MC_RGM", "RTC", "DMAMUX_1", "DMAMUX_0"],
                    ["TMU", "Block94", "LPCMP_1", "LPCMP_0", "SAI_0", "Block90", "LPSPI_3", "LPSPI_2", "LPSPI_1", "LPSPI_0", "LPI2C_1", "LPI2C_0",
                    "SIUL_VIRTWRAPPER_PDAC5", "SIUL_VIRTWRAPPER_PDAC5", "LPUART_7", "LPUART_6", "LPUART_5", "LPUART_4", "LPUART_3", "LPUART_2", "LPUART_1",
                    "LPUART_0", "FlexIO", "Block72", "Block71", "FlexCAN5", "FlexCAN4", "FlexCAN3", "FlexCAN2", "FlexCAN1", "FlexCAN0", "PIT3"],
                    ["AES_Application_2", "AES_Application_2", "AES_Application_2", "AES_Application_2", "AES_Application_1", "AES_Application_1",
                    "AES_Application_1", "AES_Application_1", "AES_Application_0", "AES_Application_0", "AES_Application_0", "AES_Application_0",
                    "AES_Accelerator", "AES_Accelerator", "AES_Accelerator", "AES_Accelerator", "Block111", "Block110",
                    "Block109", "SELFTEST_GPR", "Block107", "Block106", "Block105", "STCU", "Configuartion_GPR",
                    "Block102", "JDC", "Block100", "MU_0", "Block98", "FCCU", "CRC"],
                ][blockIndex]), value=value, sheet=sheet, start_row=row + 1, start_column=1)
            else:
                row -= 1
