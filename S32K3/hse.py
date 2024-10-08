from typing import List, Tuple

import openpyxl

from BitItem import BitItem
from IModule import IModule
from Parse import ProcessArray


class HSE(IModule):
    def __init__(self, name: str) -> None:
        self.Name = name

    def GetModuleName(self) -> str:
        return self.Name

    def UpdateSheet(self, data: List[Tuple[str, str]], sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        row = 1
        utest_base = 0x1B000000
        dcm_stat_addr = 0x402AC000
        MU_0_addr = 0x4038C000
        Cfg_GPR_addr = 0x4039C000
        for item in data:
            row += 1
            addr = int(item[0], 16)
            value = int(item[1], 16)
            if (addr == utest_base):
                sheet.cell(row=row, column=1, value="UTEST_HEADER1")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="HSE Enabled" if value == 0xDDCCBBAA else "HSE Disabled")
            elif (addr == utest_base + 0x04):
                sheet.cell(row=row, column=1, value="UTEST_HEADER2")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="HSE Enabled" if value == 0xAABBCCDD else "HSE Disabled")
            elif (addr == Cfg_GPR_addr + 0x28):
                sheet.cell(row=row, column=1, value="HSE_GPR3")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="")
                row = row + ProcessArray(array=[
                    BitItem(name="Bit31", offset=31, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit30", offset=30, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit29", offset=29, true_meaning="App should not access flash block 4", false_meaning="App can access flash block 4"),
                    BitItem(name="Bit28", offset=28, true_meaning="App should not access flash block 3", false_meaning="App can access flash block 3"),
                    BitItem(name="Bit27", offset=27, true_meaning="App should not access flash block 2", false_meaning="App can access flash block 2"),
                    BitItem(name="Bit26", offset=26, true_meaning="App should not access flash block 1", false_meaning="App can access flash block 1"),
                    BitItem(name="Bit25", offset=25, true_meaning="App should not access flash block 0", false_meaning="App can access flash block 0"),
                    BitItem(name="Bit24", offset=24, true_meaning="App should not access UTEST", false_meaning="App can access flash UTEST"),
                    BitItem(name="Bit23", offset=23, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit22", offset=22, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit21", offset=21, true_meaning="App should not erase flash block 4", false_meaning="App can erase flash block 4"),
                    BitItem(name="Bit20", offset=20, true_meaning="App should not erase flash block 3", false_meaning="App can erase flash block 3"),
                    BitItem(name="Bit19", offset=19, true_meaning="App should not erase flash block 2", false_meaning="App can erase flash block 2"),
                    BitItem(name="Bit18", offset=18, true_meaning="App should not erase flash block 1", false_meaning="App can erase flash block 1"),
                    BitItem(name="Bit17", offset=17, true_meaning="App should not erase flash block 0", false_meaning="App can erase flash block 0"),
                    BitItem(name="Bit16", offset=16, true_meaning="App should not erase UTEST", false_meaning="App can erase UTEST"),
                    BitItem(name="Bit15", offset=15, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit14", offset=14, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit13", offset=13, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit12", offset=12, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit11", offset=11, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit10", offset=10, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit9", offset=9, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit8", offset=8, true_meaning="Reserved", false_meaning=""),
                    BitItem(name="Bit7", offset=7, true_meaning="SBAF verifies IVT recovery image with random IV", false_meaning="SBAF verifies IVT recovery image with fixed IV"),
                    BitItem(name="Bit6", offset=6, true_meaning="Reserved for HSE", false_meaning="Reserved for HSE"),
                    BitItem(name="Bit5", offset=5, true_meaning="App core boot in Recovery mode by SBAF", false_meaning=""),
                    BitItem(name="Bit4", offset=4, true_meaning="SBAF Handshake erased HSE Firmware", false_meaning=""),
                    BitItem(name="Bit3", offset=3, true_meaning="SBAF Handshake erased HSE Firmware in data flash", false_meaning=""),
                    BitItem(name="Bit2", offset=2, true_meaning="SBAF Handshake erased HSE Firmware in code flash", false_meaning=""),
                    BitItem(name="Bit1", offset=1, true_meaning="MU_IF is enabled for install HSE Firmware", false_meaning=""),
                    BitItem(name="Bit0", offset=0, true_meaning="HSE FW present SBAF boot HSE FW", false_meaning="No HSE FW"),
                    ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == MU_0_addr + 0x104):
                sheet.cell(row=row, column=1, value="MU_FSR")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="")
                row = row + ProcessArray(array=[
                    BitItem(name="Bit31", offset=31, true_meaning="Reserved For Future", false_meaning=""),
                    BitItem(name="Bit30", offset=30, true_meaning="PUBLISH_NVM_RAM_TO_FLASH", false_meaning=""),
                    BitItem(name="Bit29", offset=29, true_meaning="FW_UPDATE_IN_PROGRESS", false_meaning=""),
                    BitItem(name="Bit28", offset=28, true_meaning="OEM_SU", false_meaning=""),
                    BitItem(name="Bit27", offset=27, true_meaning="CUST_SU", false_meaning=""),
                    BitItem(name="Bit26", offset=26, true_meaning="BOOT_OK", false_meaning=""),
                    BitItem(name="Bit25", offset=25, true_meaning="INSTALL_OK", false_meaning=""),
                    BitItem(name="Bit24", offset=24, true_meaning="INIT_OK", false_meaning=""),
                    BitItem(name="Bit23", offset=23, true_meaning="HSE_DEBUG_ACTIVE", false_meaning=""),
                    BitItem(name="Bit22", offset=22, true_meaning="HOST_DEBUG_ACTIVE", false_meaning=""),
                    BitItem(name="Bit21", offset=21, true_meaning="RNG_INIT_OK", false_meaning=""),
                    BitItem(name="Bit20", offset=20, true_meaning="SHE_SECURE_BOOT_OK", false_meaning=""),
                    BitItem(name="Bit19", offset=19, true_meaning="SHE_SECURE_BOOT_FINISHED", false_meaning=""),
                    BitItem(name="Bit18", offset=18, true_meaning="SHE_SECURE_BOOT_INIT", false_meaning=""),
                    BitItem(name="Bit17", offset=17, true_meaning="SHE_SECURE_BOOT", false_meaning=""),
                    BitItem(name="Bit16", offset=16, true_meaning="Reserved For Future", false_meaning=""),
                    BitItem(name="Bit15", offset=15, true_meaning="Service channel 15 in progress", false_meaning=""),
                    BitItem(name="Bit14", offset=14, true_meaning="Service channel 14 in progress", false_meaning=""),
                    BitItem(name="Bit13", offset=13, true_meaning="Service channel 13 in progress", false_meaning=""),
                    BitItem(name="Bit12", offset=12, true_meaning="Service channel 12 in progress", false_meaning=""),
                    BitItem(name="Bit11", offset=11, true_meaning="Service channel 11 in progress", false_meaning=""),
                    BitItem(name="Bit10", offset=10, true_meaning="Service channel 10 in progress", false_meaning=""),
                    BitItem(name="Bit9", offset=9, true_meaning="Service channel 9 in progress", false_meaning=""),
                    BitItem(name="Bit8", offset=8, true_meaning="Service channel 8 in progress", false_meaning=""),
                    BitItem(name="Bit7", offset=7, true_meaning="Service channel 7 in progress", false_meaning=""),
                    BitItem(name="Bit6", offset=6, true_meaning="Service channel 6 in progress", false_meaning=""),
                    BitItem(name="Bit5", offset=5, true_meaning="Service channel 5 in progress", false_meaning=""),
                    BitItem(name="Bit4", offset=4, true_meaning="Service channel 4 in progress", false_meaning=""),
                    BitItem(name="Bit3", offset=3, true_meaning="Service channel 3 in progress", false_meaning=""),
                    BitItem(name="Bit2", offset=2, true_meaning="Service channel 2 in progress", false_meaning=""),
                    BitItem(name="Bit1", offset=1, true_meaning="Service channel 1 in progress", false_meaning=""),
                    BitItem(name="Bit0", offset=0, true_meaning="Service channel 0 in progress", false_meaning=""),
                    ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == dcm_stat_addr):
                sheet.cell(row=row, column=1, value="DCM_STAT")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="")
                row = row + ProcessArray(array=[
                    BitItem(name="DCMDONE",offset=0, true_meaning="DCM Scanning Done", false_meaning="DCM Scanning Running")
                    ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
                if (value & 0x1):
                    row = row + ProcessArray(array=[
                        BitItem(name="DCMOTAA_EX", offset=18, true_meaning="Partial ABSWAP Active", false_meaning="Partial ABSWAP Inactive"),
                        BitItem(name="DCMOTAR",offset=17, true_meaning="ABSWAP High Address", false_meaning="ABSWAP Low Address"),
                        BitItem(name="DCMOTAA", offset=16, true_meaning="ABSWAP Active", false_meaning="ABSWAP Inactive"),
                        BitItem(name="DCMDBGPS", offset=10, true_meaning="Debug Password Success", false_meaning="Debug Password Error"),
                        BitItem(name="DCMOTAS", offset=9, true_meaning="ABSWAP Success", false_meaning="ABSWAP Error"),
                        BitItem(name="DCMUTS", offset=8, true_meaning="UTEST Success", false_meaning="UTEST Error"),
                        BitItem(name="DCMLCST", offset=4, true_meaning="LC Success", false_meaning="LC Error"),
                        BitItem(name="DCMERR", offset=1, true_meaning="DCM Error", false_meaning="DCM Success"),
                        ], value=value, sheet=sheet, start_row=row + 1, start_column=1)
            elif (addr == Cfg_GPR_addr + 0x38):
                sheet.cell(row=row, column=1, value="CONFIG_RAMPR")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="")
            elif (addr == Cfg_GPR_addr + 0x3C):
                sheet.cell(row=row, column=1, value="CONFIG_CFPRL")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="")
            elif (addr == Cfg_GPR_addr + 0x40):
                sheet.cell(row=row, column=1, value="CONFIG_CFPRH")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="")
            elif (addr == Cfg_GPR_addr + 0x44):
                sheet.cell(row=row, column=1, value="CONFIG_DFPR")
                sheet.cell(row=row, column=2, value=f"{value:08X}")
                sheet.cell(row=row, column=3, value="")
            else:
                row -= 1
